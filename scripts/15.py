#!/usr/bin/env python3

# Это ОЖЗ минус текущие остатки и товары в пути

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Constants
OZZ_FILE = "ОЖЗ.xlsx"
RESTOCK_FILE_TEMPLATE = "ОСТАТКИ СЮДА/остатки-{date}.xlsx"
OUTPUT_FILE = "Можно-поставить.xlsx"
DATE_FORMAT = "%d.%m.%Y"

# Step 1: Load OZZ file
if not os.path.exists(OZZ_FILE):
    raise FileNotFoundError(f"Файл с желаемым состоянием склада не найден: {OZZ_FILE}")

ozz_df = pd.read_excel(OZZ_FILE, usecols=["Название склада", "Артикул", "Общий желаемый запас", "Статус", "Рейтинг склада"])

# Step 2: Load Restock file
current_date = pd.Timestamp.now().strftime(DATE_FORMAT)
restock_file = RESTOCK_FILE_TEMPLATE.format(date=current_date)

if not os.path.exists(restock_file):
    raise FileNotFoundError(f"Файл с остатками не найден: {restock_file}")

restock_df = pd.read_excel(restock_file, skiprows=3)

# Ensure necessary columns are present
required_columns = {"Название склада", "Артикул", "Товары в пути", "Доступный к продаже товар"}
if not required_columns.issubset(restock_df.columns):
    raise ValueError(f"Файл {restock_file} должен содержать колонки: {required_columns}")

# Step 3: Merge data
merged_df = pd.merge(
    ozz_df, restock_df, on=["Название склада", "Артикул"], how="left"
)

# Step 4: Calculate "Требуется поставить"
merged_df["Товары в пути"] = merged_df["Товары в пути"].fillna(0)
merged_df["Доступный к продаже товар"] = merged_df["Доступный к продаже товар"].fillna(0)
merged_df["Требуется поставить"] = (
    merged_df["Общий желаемый запас"] - merged_df["Товары в пути"] - merged_df["Доступный к продаже товар"]
).clip(lower=0)

# Step 5: Sort by warehouse and add empty rows between warehouses
merged_df.sort_values(by="Название склада", inplace=True)
output_rows = []
for warehouse, group in merged_df.groupby("Название склада"):
    output_rows.append(group)
    output_rows.append(pd.DataFrame([{}]))  # Empty row between groups
final_df = pd.concat(output_rows, ignore_index=True)

# Step 6: Apply colors based on "Статус"
def apply_color_to_workbook(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Розовый
    green_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")  # Салатовый

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        status_cell = row[3]  # Assuming "Статус" is in the 4th column
        if status_cell.value == "Гипотеза":
            for cell in row:
                cell.fill = pink_fill
        elif status_cell.value == "Статистика":
            for cell in row:
                cell.fill = green_fill

    wb.save(file_path)

# Step 7: Save to Excel
output_columns = [
    "Название склада", "Артикул", "Общий желаемый запас", "Товары в пути", 
    "Доступный к продаже товар", "Требуется поставить", "Статус", "Рейтинг склада"
]

final_df.to_excel(OUTPUT_FILE, index=False, columns=output_columns)

# Apply colors
apply_color_to_workbook(OUTPUT_FILE)

print(f"Файл 'Черновик поставки' успешно сформирован и сохранён по пути: {OUTPUT_FILE}")

