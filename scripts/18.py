#!/usr/bin/env python3

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import shutil

# Constants
PRIORITY_FILE = "Приоритет.xlsx"
SUPPLY_FILE = "Можно-поставить.xlsx"
ITEM_RATINGS_FILE = "Рейтинг-товаров.xlsx"
OUTPUT_FOLDER = "./"
ARCHIVE_FOLDER = "АРХИВ ПОСТАВОК"

# Step 1: Load priority file
if not os.path.exists(PRIORITY_FILE):
    raise FileNotFoundError(f"Файл {PRIORITY_FILE} не найден.")

priority_df = pd.read_excel(PRIORITY_FILE)

if "Название склада" not in priority_df.columns:
    raise ValueError("Файл Приоритет.xlsx должен содержать колонку 'Название склада'.")

# Get the sorted list of warehouses by priority
warehouses = priority_df["Название склада"].tolist()

# Step 2: Prepare output file name and archive old drafts
current_date = datetime.now().strftime("%d.%m.%Y")
output_file = f"Черновик-поставки-{current_date}.xlsx"

if not os.path.exists(ARCHIVE_FOLDER):
    os.makedirs(ARCHIVE_FOLDER)

for file in os.listdir(OUTPUT_FOLDER):
    if file.startswith("Черновик-поставки") and file.endswith(".xlsx") and file != output_file:
        shutil.move(os.path.join(OUTPUT_FOLDER, file), os.path.join(ARCHIVE_FOLDER, file))

# Step 3: Load supply and ratings files
if not os.path.exists(SUPPLY_FILE):
    raise FileNotFoundError(f"Файл {SUPPLY_FILE} не найден.")
if not os.path.exists(ITEM_RATINGS_FILE):
    raise FileNotFoundError(f"Файл {ITEM_RATINGS_FILE} не найден.")

supply_df = pd.read_excel(SUPPLY_FILE)
item_ratings_df = pd.read_excel(ITEM_RATINGS_FILE)

# Ensure necessary columns exist
required_supply_columns = {"Название склада", "Артикул", "Требуется поставить", "Статус"}
if not required_supply_columns.issubset(supply_df.columns):
    raise ValueError(f"Файл {SUPPLY_FILE} должен содержать колонки: {required_supply_columns}")

if not {"Артикул", "Рейтинг товара"}.issubset(item_ratings_df.columns):
    raise ValueError("Файл Рейтинг-товаров.xlsx должен содержать колонки 'Артикул' и 'Рейтинг товара'.")

# Step 4: Create workbook and process each warehouse
wb = Workbook()
del wb[wb.active.title]  # Remove default sheet

for warehouse in warehouses:
    # Filter data for the current warehouse
    warehouse_data = supply_df[supply_df["Название склада"] == warehouse]

    # Merge with item ratings
    warehouse_data = pd.merge(
        warehouse_data,
        item_ratings_df,
        on="Артикул",
        how="left",
        suffixes=("_supply", "_item")  # Clear suffixes to avoid confusion
    )

    # Rename columns explicitly
    warehouse_data.rename(
        columns={
            "Название склада_supply": "Название склада",
            "Статус_supply": "Статус",
        },
        inplace=True,
    )

    # Select only the necessary columns after renaming
    warehouse_data = warehouse_data[[
        "Артикул",
        "Требуется поставить",
        "Статус",
        "Рейтинг товара"
    ]]

    # Fill missing ratings with 0
    warehouse_data["Рейтинг товара"] = warehouse_data["Рейтинг товара"].fillna(0)

    # Sort by item rating
    warehouse_data.sort_values(by="Рейтинг товара", ascending=False, inplace=True)

    # Create a new sheet for the warehouse
    ws = wb.create_sheet(title=warehouse[:31])  # Sheet name max length is 31 characters

    # Write headers
    headers = ["Артикул", "Требуется поставить", "Статус", "Рейтинг товара"]
    ws.append(headers)

    # Define styles
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Light orange

    # Write data
    for _, row in warehouse_data.iterrows():
        ws.append([row["Артикул"], row["Требуется поставить"], row["Статус"], row["Рейтинг товара"]])
        last_row = ws.max_row

        # Apply colors based on status
        if row["Статус"] == "Статистика":
            for cell in ws[last_row]:
                cell.fill = green_fill
        elif row["Статус"] == "Гипотеза":
            for cell in ws[last_row]:
                cell.fill = orange_fill

# Step 5: Save workbook
wb.save(output_file)
print(f"Файл '{output_file}' успешно сформирован.")

